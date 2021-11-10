import openpyxl as xl

from concurrent.futures import ThreadPoolExecutor
import os
import tempfile

from tornado import gen


#from sqlalchemy.orm import joinedload
from sqlalchemy.orm.session import object_session
#from tornado.escape import json_encode
import tornado.web

import base_handler
import errors
import model
from tornado.escape import json_encode

#import requests
#qnode_id = '9eaf4f51-e760-482c-86f8-df7028ccae51'
#survey_id = '60f224d0-a96c-41a1-9a46-3fa4aed86262'
# table first  column number, user to check if template match survey

table1FirstColumn = 1
table2FirstColumn = 12
targetColumnName = 'Target'
BUF_SIZE = 4096
MAX_WORKERS = 4
class ExportAssetHandler(base_handler.BaseHandler):
    executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)
        


    @tornado.web.authenticated
    @gen.coroutine
    def get(self, submission_id, survey_id, program_id, path, uri):
        son={}
        #path = os.getcwd() + "/doc/"
        #path = 'src/app/client/'
        path = path[1:len(path)-1]
        uri = uri[1:len(uri)-1]
        with model.session_scope() as session:
            #user_session = self.get_user_session(session)
            survey = (
                session.query(model.Survey)
                    .filter(model.Survey.id == survey_id)
                    .filter(model.Survey.program_id == program_id)
                    .filter(model.Survey.deleted != True).first())     
            if survey: 
                surveyName = survey.title 
                templateFile = path + uri + surveyName + ' Template.xlsx'
                reportName = uri + surveyName + ' Report.xlsx'
                #r = requests.get('http://EITWKS49:8081'+'/' + 'AssetManagement.xlsx', allow_redirects=True)
                #open('http://EITWKS49:8081'+'/'+surveyName + 'AssetManagement.xlsx', 'wb').write(r.content)
                if os.path.isfile(templateFile):
                    file = open(templateFile, 'rb')
                    template = xl.load_workbook(file)     
                    tSheet = template.worksheets[0]                                  
                    #check survey template
                    if tSheet.cell(1,1).value == surveyName + " RAW DATA":
                        son = self.export(submission_id, survey_id, program_id, template, tSheet, path, reportName)
                        #son["report"] = reportName
                    else:
                        son["errorMessage"] = "No template to create " + surveyName + " report"
                else:
                    son["errorMessage"] = "The template file for " + surveyName + " report not exist"
            else:
                son["errorMessage"] = "Not suvery to create asset management report"
        #son["path"] = path
        #son["uri"] = uri
        self.set_header("Content-Type", "application/json")
        self.write(json_encode(son))
        self.finish()


    def export(self, submission_id, survey_id, program_id, template, tSheet, path, reportName):
        fileName= path + reportName
        result= {}
        # if return -1, no column in template file
        targetColumn = self.getColumnNumber(targetColumnName,tSheet)
        with model.session_scope() as session:
            #user_session = self.get_user_session(session)
            qnodes = (session.query(model.QuestionNode)
                    #.filter((model.QuestionNode.parent_id == qnode_id) | (model.QuestionNode.survey_id == survey_id)) 
                    .filter(model.QuestionNode.survey_id == survey_id)
                    .filter(model.QuestionNode.program_id == program_id)
                    .filter(model.QuestionNode.deleted != True )
                    .order_by(model.QuestionNode.seq))
      
            sheets = []
            targets = []
            for q in qnodes:   

                qTarget = (
                    session.query(model.ResponseNode)
                        .filter(model.ResponseNode.qnode_id == q.id)
                        .filter(model.ResponseNode.submission_id == submission_id).first())
                # targetColumn < 0 , no target column in template file, so should not fill valumn to any cell
                if (qTarget is not None and qTarget.urgency is not None):
                    targets.append({table2FirstColumn: q.seq + 1, (1+table2FirstColumn) : str(q.seq + 1) + ' ' + q.title, (targetColumn ): qTarget.urgency})
                else: 
                    targets.append({table2FirstColumn: q.seq + 1, (1+table2FirstColumn) : str(q.seq + 1) + ' ' + q.title, (targetColumn ): 0}) 
                answerResponses = (
                    session.query(model.Measure, model.QnodeMeasure)
                        .filter(model.Measure.deleted != True)
                        .filter(model.Measure.program_id == program_id)
                        .filter(model.QnodeMeasure.program_id == model.Measure.program_id)
                        .filter(model.Measure.id == model.QnodeMeasure.measure_id)
                        .filter(model.QnodeMeasure.qnode_id == q.id ) 
                        #.filter(model.QuestionMeasure.program_id == model.Program.id)
                        #.filter(model.Program.deleted != True)
                        .order_by(model.QnodeMeasure.seq))
 
                #row = {"subject_head": q.group, "subject_number" : q.seq + 1, 'subject_name': q.title}
                for response in answerResponses:
                    #subMeasure = (
                    #    session.query(model.Measure)
                    #    .filter(model.Measure.measure_id == response.Measure.id)
                    #    #.filter(model.Measure.response_type_id == response.Measure.response_type_id)
                    #    .filter(model.Measure.submeasure_seq > 0)
                    #    .order_by(model.Measure.submeasure_seq))
                    #if (survey_id == 'c9255e07-96eb-4772-8606-7a78e549ce1e'):  
                    # get first submeasure depend on measure_id and program_id, 
                    subMeasure = (
                        session.query(model.Measure)
                        .filter(model.Measure.measure_id == response.Measure.id)
                        .filter(model.Measure.program_id == program_id)
                        .filter(model.Measure.submeasure_seq == 1)
                        .filter(model.Measure.deleted != True))
                    if (len(subMeasure.all()) > 0): 
                        # no submeasure depend on measure_id and program_id, not consider program 
                        subMeasure = (
                            session.query(model.Measure)
                            .filter(model.Measure.measure_id == response.Measure.id)
                            #.filter(model.Measure.response_type_id == response.Measure.response_type_id)
                            .filter(model.Measure.program_id == program_id)
                            .filter(model.Measure.submeasure_seq > 0)
                            .filter(model.Measure.deleted != True)
                            .order_by(model.Measure.submeasure_seq))
                    else:
                        # exist submeasure depend on measure_id and program_id, need consider program
                        subMeasure = (
                            session.query(model.Measure)
                            .filter(model.Measure.measure_id == response.Measure.id)
                            #.filter(model.Measure.response_type_id == response.Measure.response_type_id)
                            #.filter(model.Measure.program_id == program_id)
                            .filter(model.Measure.submeasure_seq > 0)
                            .filter(model.Measure.deleted != True)
                            .order_by(model.Measure.submeasure_seq))                           
                    answerResponse = (
                        session.query(model.Response, model.Measure)
                            .filter(model.Response.measure_id == model.Measure.id)
                            .filter(model.Response.submission_id == submission_id)
                            .filter(model.Measure.id == response.Measure.id)
                            .first()                  
                    )

                    for r, p in enumerate(subMeasure):
                        #if answerResponse.Response.variables != {}:
                        score = 0
                        if answerResponse and len(answerResponse.Response.response_parts)>0:
                            #calculate sub measure score

                            rt = answerResponse.Measure.response_type.parts                    
                            weight = p.weight

                            if (weight is None or weight == ""):
                                weight = 0
                            if p.submeasure_seq is None:
                                if (not (answerResponse.Response.response_parts[r] is None or answerResponse.Response.response_parts[r] == {})):
                                    if ('index' in answerResponse.Response.response_parts[r].keys()):
                                        score = weight * answerResponse.Response.response_parts[r]['index']
                                    if ('value' in answerResponse.Response.response_parts[r].keys()):
                                        score = weight * answerResponse.Response.response_parts[r]['value']   
                            else:           
                                for j, t in enumerate(rt):
                                    if (t['submeasure_seq'] ==  p.submeasure_seq ):
                                        if (not (answerResponse.Response.response_parts[j] is None or answerResponse.Response.response_parts[j] == {})):
                                            if ('index' in answerResponse.Response.response_parts[j].keys()):
                                                score += weight * answerResponse.Response.response_parts[j]['index']
                                            if ('value' in answerResponse.Response.response_parts[j].keys()):
                                                score += weight * answerResponse.Response.response_parts[j]['value'] 
                             
                 
                        sheets.append(
                            { 
                                table1FirstColumn: q.group, 
                                (1+table1FirstColumn): q.seq + 1, 
                                (2+table1FirstColumn): q.title,
                                (3+table1FirstColumn): response.QnodeMeasure.seq + 1, 
                                (4+table1FirstColumn): response.Measure.title,
                                (5+table1FirstColumn): str(q.seq + 1) + '.' + str(response.QnodeMeasure.seq + 1) +'.' + str(p.submeasure_seq),
                                (6+table1FirstColumn): p.description, 
                                (7+table1FirstColumn): score
                            })  
        

                        #for r, p in enumerate(rt):
                        #    if (False and not 'submeasure_seq' in p ):
                        #        if (seq > 0):
                        #            question += 1
                        #            if (hasAnswer):
                        #                answer += 1
                        #            else:
                        #                hasAnswer=True
                        #        if ((answerResponse.Response.response_parts[r] is None or answerResponse.Response.response_parts[r] == {} or
                        #            ((not 'index' in answerResponse.Response.response_parts[r].keys()) and 
                        #            (not 'value' in answerResponse.Response.response_parts[r].keys()))) and hasAnswer):
                        #            hasAnswer=False    
                        #    else:           
                        #        if (False and (hasAnswer or seq != p['submeasure_seq'])):
                        #            if (seq != p['submeasure_seq']):
                        #                if (seq > 0):
                        #                    question += 1
                        #                    if (hasAnswer):
                        #                        answer += 1
                        #                    else:
                        #                        hasAnswer=True

                        #                seq = p['submeasure_seq']
                          
                        #            if ((answerResponse.Response.response_parts[r] is None or answerResponse.Response.response_parts[r] == {} or
                        #                ((not 'index' in answerResponse.Response.response_parts[r].keys()) and 
                        #                (not 'value' in answerResponse.Response.response_parts[r].keys()))) and hasAnswer):
                        #                    hasAnswer=False        
            sRows = len(sheets)
            tRows = len(targets)
            if (tSheet.cell(sRows+2,table1FirstColumn).value and tSheet.cell(sRows+3,table1FirstColumn).value is None) and (tSheet.cell(tRows+2,table2FirstColumn).value and tSheet.cell(tRows+3,table2FirstColumn).value is None):
                # saving the destination excel file 
                for j, r in enumerate(sheets):
                    for i in range(1,len(r.keys())+1):
                        tSheet.cell(3+j,i).value=r[i]
                for j, r in enumerate(targets):  
                    keys= list(r.keys())
                    for k in keys:
                        # if key < 0 , value should not fill to cell 
                        if ( k > -1 ) :
                            tSheet.cell(3+j,k).value=r[k] 
                if os.path.isfile(fileName):    
                    os.remove(fileName) 
                template.save(fileName)  
                template.close()
                result["message"] = "Export finished"
                result["report"] = reportName
            else:
                result["errorMessage"] = "Template file not match this survey to export asset management report"
        return result

    # find column index  in template file, no this column in template file, return -1
    def getColumnNumber(self, columnName, sheet):
        index = table2FirstColumn
        while index == index:           
                if sheet.cell(2,index).value == columnName:
                    return index
                if sheet.cell(2,index).value == None:
                    return -1
                index = index +1    

        return -1


 
       
  